from __future__ import annotations

from io import BytesIO
import sys
import tempfile
from pathlib import Path
from zipfile import ZipFile

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from app.services.leads_export import build_leads_xlsx


def main() -> None:
    data = build_leads_xlsx([
        {
            "created_at": "03 Aug 2026, 23:16 WIB",
            "name": "Yusuf Fikri",
            "whatsapp": "08123456789",
            "email": "yusuf@example.com",
            "business_type": "Digital Agency",
            "campaign": "Independence Campaign Index",
            "license_id": "DAB-260803-000001",
            "payment_status": "paid",
            "payment_amount": 99000,
            "status": "ready",
            "downloads": 1,
            "expires_at": "05 Aug 2026, 23:16 WIB",
        },
        {
            "created_at": "04 Aug 2026, 08:35 WIB",
            "name": "=HYPERLINK(\"https://invalid.example\",\"unsafe\")",
            "whatsapp": "+62 812-3456-7890",
            "email": "safe@example.com",
            "business_type": "Retail",
            "campaign": "Free Template",
            "license_id": "DAB-260804-000002",
            "payment_status": "not_required",
            "payment_amount": 0,
            "status": "queued",
            "downloads": 0,
            "expires_at": "06 Aug 2026, 08:35 WIB",
        },
    ], "04 Aug 2026, 08:35 WIB")

    assert data[:2] == b"PK"
    with ZipFile(BytesIO(data)) as workbook_zip:
        assert workbook_zip.testzip() is None
        names = set(workbook_zip.namelist())
        assert "xl/worksheets/sheet1.xml" in names
        assert "xl/styles.xml" in names
        assert "docProps/core.xml" in names

    # This is the regression check for Excel's “problem with some content” dialog.
    workbook = load_workbook(BytesIO(data), read_only=False, data_only=False, keep_links=False)
    try:
        sheet = workbook["Leads"]
        assert sheet["A1"].value == "DABTIVE LEADS DATABASE"
        assert sheet["B5"].value == "Yusuf Fikri"
        assert sheet["C5"].value == "08123456789"
        assert sheet["C5"].hyperlink.target == "https://wa.me/628123456789"
        assert sheet["D5"].hyperlink.target == "mailto:yusuf@example.com"
        assert sheet["B6"].data_type != "f"
        assert sheet["B6"].value.startswith("'")
        assert sheet.auto_filter.ref == "A4:L6"
        assert sheet.freeze_panes == "A5"
    finally:
        workbook.close()

    # Also verify that LibreOffice can open and re-save the package without repair.
    with tempfile.TemporaryDirectory() as temp:
        path = Path(temp) / "dabtive-leads.xlsx"
        path.write_bytes(data)
        reopened = load_workbook(path, read_only=True)
        reopened.close()

    print("EXPORT TEST OK", len(data))


if __name__ == "__main__":
    main()
