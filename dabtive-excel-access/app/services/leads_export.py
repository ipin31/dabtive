from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import re

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_PHONE_DIGITS_RE = re.compile(r"\D+")


def _safe_text(value: object) -> str:
    """Return Excel-safe text without turning user input into a formula."""
    text = "" if value is None else str(value)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    if text.startswith(_FORMULA_PREFIXES):
        return "'" + text
    return text


def _whatsapp_url(value: object) -> str | None:
    digits = _PHONE_DIGITS_RE.sub("", "" if value is None else str(value))
    if not digits:
        return None
    if digits.startswith("0"):
        digits = "62" + digits[1:]
    elif not digits.startswith("62"):
        digits = "62" + digits
    return f"https://wa.me/{digits}"


def _email_url(value: object) -> str | None:
    text = "" if value is None else str(value).strip()
    if not text or "@" not in text or any(char in text for char in "\r\n"):
        return None
    return f"mailto:{text}"


def build_leads_xlsx(rows: list[dict[str, object]], generated_at: str) -> bytes:
    """Build a native, validated XLSX export with OpenPyXL.

    Earlier versions hand-crafted the OOXML ZIP. Although readable by some
    spreadsheet programs, desktop Excel could show a repair warning. This
    implementation delegates the complete package structure to OpenPyXL and
    validates the saved workbook before returning it.
    """
    headers = [
        "Tanggal Request",
        "Nama",
        "Nomor WhatsApp",
        "Email",
        "Jenis Bisnis",
        "File / Campaign",
        "License ID",
        "Status Pembayaran",
        "Nominal",
        "Status File",
        "Jumlah Download",
        "Link Expired",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    wb.properties.title = "Dabtive Leads Database"
    wb.properties.subject = "Daftar leads dan requester file"
    wb.properties.creator = "Dabtive Excel Access"
    wb.properties.lastModifiedBy = "Dabtive Excel Access"
    now = datetime.now(timezone.utc).replace(microsecond=0)
    wb.properties.created = now
    wb.properties.modified = now

    black = "111111"
    red = "E1162B"
    white = "FFFFFF"
    border_color = "E1E1E5"
    green_fill = "E8F8EF"
    red_fill = "FFE7EA"
    amber_fill = "FFF3CF"

    thin = Side(style="thin", color=border_color)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor=black)
    header_fill = PatternFill("solid", fgColor=red)
    success_fill = PatternFill("solid", fgColor=green_fill)
    error_fill = PatternFill("solid", fgColor=red_fill)
    pending_fill = PatternFill("solid", fgColor=amber_fill)

    ws.merge_cells("A1:L1")
    ws["A1"] = "DABTIVE LEADS DATABASE"
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=white)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Generated: {_safe_text(generated_at)} · Total leads: {len(rows):,}"
    ws["A2"].font = Font(name="Aptos", size=10, bold=True, color=white)
    ws["A2"].fill = header_fill
    ws["A2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 8

    for column, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=column, value=header)
        cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        cell.fill = header_fill
        cell.border = cell_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 26

    for row_number, item in enumerate(rows, 5):
        status = _safe_text(item.get("status")).upper()
        payment_status = _safe_text(item.get("payment_status")).upper()
        values: list[object] = [
            _safe_text(item.get("created_at")),
            _safe_text(item.get("name")),
            _safe_text(item.get("whatsapp")),
            _safe_text(item.get("email")),
            _safe_text(item.get("business_type")),
            _safe_text(item.get("campaign")),
            _safe_text(item.get("license_id")),
            payment_status,
            int(item.get("payment_amount") or 0),
            status,
            int(item.get("downloads") or 0),
            _safe_text(item.get("expires_at")),
        ]

        for column, value in enumerate(values, 1):
            cell = ws.cell(row=row_number, column=column, value=value)
            cell.font = Font(name="Aptos", size=10, color=black)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Preserve WhatsApp as text and make it operationally useful.
        ws.cell(row=row_number, column=3).number_format = "@"
        wa_url = _whatsapp_url(item.get("whatsapp"))
        if wa_url:
            ws.cell(row=row_number, column=3).hyperlink = wa_url
            ws.cell(row=row_number, column=3).style = "Hyperlink"
            ws.cell(row=row_number, column=3).border = cell_border
            ws.cell(row=row_number, column=3).alignment = Alignment(vertical="center")

        email_url = _email_url(item.get("email"))
        if email_url:
            ws.cell(row=row_number, column=4).hyperlink = email_url
            ws.cell(row=row_number, column=4).style = "Hyperlink"
            ws.cell(row=row_number, column=4).border = cell_border
            ws.cell(row=row_number, column=4).alignment = Alignment(vertical="center")

        ws.cell(row=row_number, column=8).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_number, column=8).fill = success_fill if payment_status in {"PAID", "NOT_REQUIRED"} else pending_fill

        ws.cell(row=row_number, column=9).number_format = '[$Rp-421] #,##0'
        ws.cell(row=row_number, column=9).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_number, column=10).alignment = Alignment(horizontal="center", vertical="center")
        if status == "READY":
            ws.cell(row=row_number, column=10).fill = success_fill
        elif status == "FAILED":
            ws.cell(row=row_number, column=10).fill = error_fill
        else:
            ws.cell(row=row_number, column=10).fill = pending_fill
        ws.cell(row=row_number, column=11).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_number].height = 22

    widths = [21, 24, 19, 31, 24, 29, 23, 18, 16, 16, 17, 21]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    last_row = max(4, len(rows) + 4)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:L{last_row}"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    data = buffer.getvalue()

    # Fail closed: never send an XLSX that our parser cannot reopen.
    validation = load_workbook(BytesIO(data), read_only=True, data_only=False, keep_links=False)
    try:
        sheet = validation["Leads"]
        if sheet["A1"].value != "DABTIVE LEADS DATABASE":
            raise RuntimeError("Validasi export XLSX gagal")
    finally:
        validation.close()
    return data
